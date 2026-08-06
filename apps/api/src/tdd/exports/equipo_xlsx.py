"""Lectura y plantilla del XLSX de inventario `[REQ]` §7 / P-15.

Es la capa fina que toca `openpyxl`. Todo lo que decide algo —qué columna es
qué campo, qué fila es válida, qué se sobrescribe— vive en
`tdd.equipment.importacion`, que es lógica pura y se prueba sin ficheros.

`[LIM]` Solo se lee **la primera hoja**. Un libro con varias hojas de inventario
no se recorre entero: se avisa de cuántas hay y se importa la primera. Adivinar
cuál de cinco hojas es la buena sería justo el tipo de suposición que produce
una importación silenciosamente incompleta.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tdd.equipment.importacion import CABECERAS, ENUMERADOS

AZUL = "1D4ED8"
GRIS = "F1F5F9"


class LibroIlegible(ValueError):
    """El fichero no es un XLSX que se pueda abrir."""


@dataclass(frozen=True, slots=True)
class Hoja:
    cabeceras: list[str]
    filas: list[list[str]]
    nombre: str
    #: Cuántas hojas tenía el libro. Se enseña cuando hay más de una.
    total_hojas: int


def _texto(valor: object) -> str:
    """Una celda como cadena, sin que Excel meta ruido.

    Un año tecleado en una celda numérica llega como `2010.0`, y una fecha
    completa como `datetime`. Los dos casos pasan en hojas reales.
    """
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return str(valor.year)
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def leer(contenido: bytes) -> Hoja:
    """Lee la primera hoja del libro. No valida nada: de eso va `importacion`."""
    try:
        # `data_only`: si la celda lleva una fórmula queremos el resultado que
        # Excel dejó guardado, no «=B2*2», que no es un dato.
        libro = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl lanza de todo
        raise LibroIlegible(
            "El fichero no se puede abrir como libro de Excel (.xlsx). "
            "Si lo tiene en .xls o en .csv, guárdelo antes como .xlsx."
        ) from exc

    hoja = libro.worksheets[0]
    # El recuento y el título se leen ANTES de cerrar: un libro cerrado en modo
    # de solo lectura ya no responde a `worksheets`.
    total, nombre = len(libro.worksheets), hoja.title
    filas = [[_texto(c) for c in fila] for fila in hoja.iter_rows(values_only=True)]
    libro.close()

    if not filas:
        return Hoja(cabeceras=[], filas=[], nombre=nombre, total_hojas=total)
    return Hoja(cabeceras=filas[0], filas=filas[1:], nombre=nombre, total_hojas=total)


#: Orden de la plantilla. Es el mismo de la ficha en pantalla, para que quien
#: rellena la hoja y quien la revisa después estén mirando lo mismo.
ORDEN = (
    "asset",
    "tag",
    "equipment_type",
    "technical_system",
    "manufacturer",
    "model",
    "serial_number",
    "install_year",
    "expected_life_years",
    "condition",
    "obsolescence",
    "criticality",
    "quantity",
    "unit",
    "has_documentation",
    "notes",
)

TITULOS = {
    "asset": "Activo",
    "tag": "Etiqueta",
    "equipment_type": "Tipo de equipo",
    "technical_system": "Sistema técnico",
    "manufacturer": "Fabricante",
    "model": "Modelo",
    "serial_number": "Número de serie",
    "install_year": "Año de instalación",
    "expected_life_years": "Vida útil esperada",
    "condition": "Estado de conservación",
    "obsolescence": "Obsolescencia",
    "criticality": "Criticidad",
    "quantity": "Cantidad",
    "unit": "Unidad",
    "has_documentation": "Documentación",
    "notes": "Observaciones",
}

AYUDAS = {
    "asset": "Obligatorio. El nombre o el código de un activo YA dado de alta en el encargo",
    "tag": "Como está rotulado en la sala: CL-01, AS-Norte. Único dentro del activo",
    "equipment_type": "Obligatorio. Enfriadora, ascensor, cuadro general…",
    "technical_system": "Uno de los 14 del catálogo. Si no casa, el equipo entra sin clasificar",
    "install_year": "Va con la vida útil esperada: o los dos o ninguno",
    "expected_life_years": "En años. La vida residual NO se teclea: se calcula",
    "quantity": "Por defecto 1",
    "unit": "Por defecto «ud»",
    "has_documentation": "Sí / No",
}


def plantilla(activos: list[str], sistemas: list[str]) -> bytes:
    """El libro vacío que se descarga para rellenar.

    Lleva dentro **los activos del encargo y los 14 sistemas**, en una hoja
    aparte. Sin eso, quien rellena la hoja escribe el nombre del edificio de
    memoria y la mitad de las filas fallan al importar por una tilde.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    for i, campo in enumerate(ORDEN, 1):
        celda = ws.cell(row=1, column=i, value=TITULOS[campo])
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=AZUL)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ayuda = ws.cell(row=2, column=i, value=AYUDAS.get(campo, ""))
        ayuda.font = Font(italic=True, size=9, color="475569")
        ayuda.fill = PatternFill("solid", fgColor=GRIS)
        ayuda.alignment = Alignment(vertical="top", wrap_text=True)

        ws.column_dimensions[get_column_letter(i)].width = 22
    ws.row_dimensions[2].height = 42
    ws.freeze_panes = "A3"

    # La fila 2 es ayuda, no un dato. Se dice aquí y lo repite el importador:
    # una fila con «Obligatorio. El nombre o el código…» en la columna Activo
    # no casa con ningún activo y sale como error, así que nadie la cuela sin
    # enterarse.
    ws.cell(
        row=3,
        column=1,
        value="⟵ Borre esta fila y la de ayuda antes de importar, o déjelas: se rechazan solas",
    ).font = Font(italic=True, size=9, color="B45309")

    ref = wb.create_sheet("Valores admitidos")
    ref.column_dimensions["A"].width = 30
    ref.column_dimensions["B"].width = 40
    fila = 1

    def bloque(titulo: str, valores: list[str]) -> None:
        nonlocal fila
        celda = ref.cell(row=fila, column=1, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=AZUL)
        fila += 1
        for valor in valores:
            ref.cell(row=fila, column=1, value=valor)
            fila += 1
        fila += 1

    bloque("Activos de este encargo", activos or ["(el encargo no tiene activos todavía)"])
    bloque("Sistemas técnicos", sistemas)
    for campo, tabla in ENUMERADOS.items():
        bloque(TITULOS[campo], sorted({v.replace("_", " ").capitalize() for v in tabla.values()}))

    alias = wb.create_sheet("Cabeceras admitidas")
    alias.column_dimensions["A"].width = 26
    alias.column_dimensions["B"].width = 60
    alias.cell(row=1, column=1, value="Columna").font = Font(bold=True)
    alias.cell(row=1, column=2, value="También se reconoce como").font = Font(bold=True)
    for i, campo in enumerate(ORDEN, 2):
        alias.cell(row=i, column=1, value=TITULOS[campo])
        alias.cell(row=i, column=2, value=", ".join(CABECERAS[campo]))

    salida = BytesIO()
    wb.save(salida)
    return salida.getvalue()
