"""Las etiquetas que la plantilla CAPEX espera en cada celda, en su idioma.

**La plantilla es la fuente de verdad de sus propias etiquetas, no la base de
datos.** Las columnas «Objeto», «Zona Afectada», «Riesgo», «Concepto» y
«Recuperable a inquilino/s» tienen desplegable con lista cerrada, y de esas
listas se alimentan además las tablas dinámicas y los gráficos. Escribir
«Cuartos Técnicos» donde la plantilla inglesa espera «Technical Rooms» no da
un error: da una hoja que se abre, se ve bien y cuyos gráficos salen vacíos.

Por eso aquí no se traduce nada. Se lee la etiqueta del fichero que se va a
rellenar, indexada por el **código** del catálogo de la aplicación. Añadir un
idioma es añadir una plantilla, no tocar la base de datos.

El emparejamiento **no es posicional**, y esa es la parte que costó. La idea
obvia —«`HC.H09.02` es el segundo objeto de la fila de H09»— funciona en Hard
Costs pero miente en Medioambiental y ESG: el catálogo conserva `General` en la
primera posición para no renumerar las líneas ya codificadas cuando llegó el
desglose, mientras que la plantilla lo pone **el último** de su lista. Con un
emparejamiento por índice, un hallazgo de `MA.General.01` habría salido escrito
como «Situación legal».

Así que el puente se tiende por el **nombre español**: el catálogo sembrado
(`data/catalogos/codigos_capex.csv`) dice cómo se llama cada código, la
plantilla española dice en qué casilla vive ese nombre, y la plantilla del
idioma pedido dice qué pone en esa misma casilla. Un nombre que no aparezca en
la plantilla revienta al construir el vocabulario, no al exportar.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from functools import lru_cache
from typing import Any

from openpyxl import load_workbook

from tdd.exports.plantilla_capex import FICHERO, PLANTILLAS

#: Posición de las hojas de las que se lee, contando desde 0. Por posición y no
#: por nombre: el nombre cambia con el idioma.
POS_CATEGORIAS = 0
POS_OBJETO = 1
POS_ACTIVO = 2
POS_LEYENDA = 3

#: «00 Datos Objeto»: fila de cada categoría de nivel 2.
FILA_OBJETO: dict[str, int] = {
    **{f"HC.H{n:02d}": 2 + n for n in range(1, 16)},
    "MA.General": 20,
    "ESG.General": 22,
}
#: Primera columna de la lista de objetos, y cuántas puede haber.
COL_PRIMER_OBJETO = 3  # C
MAX_OBJETOS = 16  # C..R

#: «Leyenda»: los cinco plazos y los cuatro grados de riesgo.
FILA_PLAZO = {"CORTO": 3, "MEDIO": 4, "LARGO": 5, "MEJORAS": 6, "OTRO": 7}
FILA_RIESGO = {"04": 10, "03": 11, "02": 12, "01": 13}

#: «00 Datos Activo»: rejilla de zonas. Fila 3 lleva los tipos de edificio.
FILA_TIPOS = 3
COL_PRIMER_TIPO = 8  # H
TIPOS = 6  # H..M
FILA_PRIMERA_ZONA = 5
FILA_ULTIMA_ZONA = 26

#: «00 Datos Activo»: conceptos en P4:P13 y recuperable en Q4:Q6. El orden de
#: los conceptos **no** es alfabético ni el de sus letras: se indexa por letra.
COL_CONCEPTO = 16  # P
FILA_PRIMER_CONCEPTO = 4
COL_RECUPERABLE = 17  # Q
FILA_PRIMER_RECUPERABLE = 4

#: Código de `capex_concept` → letra con la que la plantilla lo numera.
LETRA_DE_CONCEPTO = {
    "MANTENIMIENTO": "a",
    "REPARACION": "b",
    "NORMATIVA": "c",
    "VIDA_UTIL": "d",
    "SEGURIDAD": "e",
    "MEJORA": "f",
    "OTRO": "g",
    "ESG": "h",
    "MEDIOAMBIENTAL": "i",
    "SOFT_COST": "j",
}
#: Orden en que la plantilla ofrece SI / NO / N.A.
ORDEN_RECUPERABLE = ("SI", "NO", "NA")

#: Orden de las columnas de tipo de edificio (H..M). `[SUP]` La plantilla pone
#: OFICINAS delante e INDUSTRIAL después; `asset_typology.sort_order` al revés.
COD_TIPOLOGIA = {
    0: "OFICINAS",
    1: "INDUSTRIAL",
    2: "HOTEL",
    3: "COMERCIAL",
    4: "SANITARIO",
    5: "OTROS",
}


class FaltaEnLaPlantilla(KeyError):
    """Se pidió una etiqueta que la plantilla no tiene.

    Se prefiere reventar a escribir una cadena que el desplegable no reconoce:
    una hoja con un valor fuera de lista se abre igual y engaña.
    """


@dataclass(frozen=True, slots=True)
class Vocabulario:
    """Las listas cerradas de una plantilla, ya leídas."""

    idioma: str
    #: `código de nivel 3` → etiqueta. Ej. `HC.H09.02` → «CGBT».
    objetos: dict[str, str]
    #: `(código de zona, código de tipología)` → etiqueta.
    zonas: dict[tuple[str, str], str]
    #: `código de tipología` → etiqueta del tipo de edificio.
    tipos_de_edificio: dict[str, str]
    riesgos: dict[str, str]
    plazos: dict[str, str]
    conceptos: dict[str, str]
    recuperables: dict[str, str]

    def objeto(self, codigo: str | None) -> str | None:
        if codigo is None:
            return None
        try:
            return self.objetos[codigo]
        except KeyError as e:
            raise FaltaEnLaPlantilla(
                f"la plantilla «{self.idioma}» no tiene el objeto {codigo!r}"
            ) from e

    def zona(self, codigo: str | None, tipologia: str) -> str | None:
        if codigo is None:
            return None
        etiqueta = self.zonas.get((codigo, tipologia))
        if etiqueta is None:
            raise FaltaEnLaPlantilla(
                f"la plantilla «{self.idioma}» no ofrece la zona {codigo!r} "
                f"para un edificio de tipo {tipologia!r}"
            )
        return etiqueta

    def riesgo(self, codigo: str | None) -> str | None:
        return self.riesgos.get(codigo) if codigo else None

    def concepto(self, codigo: str | None) -> str | None:
        return self.conceptos.get(codigo) if codigo else None

    def recuperable(self, valor: str | None) -> str | None:
        return self.recuperables.get(valor) if valor else None

    def tipo_de_edificio(self, tipologia: str) -> str | None:
        return self.tipos_de_edificio.get(tipologia)


def _texto(celda: Any) -> str | None:
    v = celda.value
    if v is None:
        return None
    v = str(v).strip()
    return None if v in ("", "-") else v


@lru_cache(maxsize=4)
def leer(idioma: str = "es") -> Vocabulario:
    """Lee las listas de la plantilla del idioma pedido.

    Se lee con `openpyxl` y **no se guarda nada**: la destrucción de partes que
    obliga a rellenar por XML solo ocurre al escribir.
    """
    if idioma not in FICHERO:
        raise ValueError(f"idioma no soportado: {idioma!r}")
    libro = load_workbook(PLANTILLAS / FICHERO[idioma], read_only=False, data_only=False)
    hojas = libro.worksheets
    objeto, activo, leyenda = hojas[POS_OBJETO], hojas[POS_ACTIVO], hojas[POS_LEYENDA]

    # Por nombre y no por índice: ver la cabecera del módulo. `General` ocupa
    # la primera posición del catálogo en MA y ESG y la última en la plantilla.
    objetos: dict[str, str] = {}
    for codigo, (fila, columna) in _casillas_de_objeto().items():
        etiqueta = _texto(objeto.cell(fila, columna))
        if etiqueta is not None:
            objetos[codigo] = etiqueta

    tipos_de_edificio = {
        COD_TIPOLOGIA[i]: t
        for i in range(TIPOS)
        if (t := _texto(activo.cell(FILA_TIPOS, COL_PRIMER_TIPO + i)))
    }

    # La rejilla de zonas se indexa por POSICIÓN, no por su texto: la casilla
    # que en la plantilla española dice «Cubierta» dice «Roof» en la inglesa, y
    # las dos son la zona `CUBIERTA`. Emparejar por el literal solo habría
    # funcionado en español, que es justo el idioma que no da problemas.
    zonas: dict[tuple[str, str], str] = {}
    for (fila, i), codigo in _rejilla_de_zonas().items():
        etiqueta = _texto(activo.cell(fila, COL_PRIMER_TIPO + i))
        if etiqueta:
            zonas[(codigo, COD_TIPOLOGIA[i])] = etiqueta

    riesgos = {c: _texto(leyenda.cell(f, 2)) or c for c, f in FILA_RIESGO.items()}
    plazos = {c: _texto(leyenda.cell(f, 2)) or c for c, f in FILA_PLAZO.items()}

    por_letra: dict[str, str] = {}
    for i in range(len(LETRA_DE_CONCEPTO)):
        etiqueta = _texto(activo.cell(FILA_PRIMER_CONCEPTO + i, COL_CONCEPTO))
        if etiqueta and (m := re.match(r"([a-z])\.", etiqueta)):
            por_letra[m.group(1)] = etiqueta
    conceptos = {
        codigo: por_letra[letra]
        for codigo, letra in LETRA_DE_CONCEPTO.items()
        if letra in por_letra
    }

    recuperables = {}
    for i, codigo in enumerate(ORDEN_RECUPERABLE):
        etiqueta = _texto(activo.cell(FILA_PRIMER_RECUPERABLE + i, COL_RECUPERABLE))
        if etiqueta:
            recuperables[codigo] = etiqueta

    libro.close()
    return Vocabulario(
        idioma=idioma,
        objetos=objetos,
        zonas=zonas,
        tipos_de_edificio=tipos_de_edificio,
        riesgos=riesgos,
        plazos=plazos,
        conceptos=conceptos,
        recuperables=recuperables,
    )


def _normalizar(texto: str) -> str:
    """Para comparar nombres entre el catálogo y la plantilla sin tropezar con
    una tilde de más, un espacio final o una mayúscula."""
    limpio = " ".join(texto.split()).lower()
    return limpio.translate(str.maketrans("áéíóúüàèìòù", "aeiouuaeiou"))


@lru_cache(maxsize=1)
def _nombres_del_catalogo() -> dict[str, tuple[str, str]]:
    """`código de nivel 3` → `(capítulo, nombre español)`, del CSV sembrado.

    El CSV es el que genera `tools/generar_catalogos.py` desde §5.3, así que es
    la misma fuente de la que sale la base de datos: si el documento y la
    plantilla se separan, se nota aquí y no en un Excel ya enviado.
    """
    import csv

    from tdd.catalogs.seeding import CATALOGOS

    nombres: dict[str, tuple[str, str]] = {}
    with (CATALOGOS / "codigos_capex.csv").open(encoding="utf-8") as f:
        for fila in csv.DictReader(f):
            if fila["level"] == "3":
                nombres[fila["code"]] = (fila["parent_code"], fila["name_es"])
    return nombres


@lru_cache(maxsize=1)
def _casillas_de_objeto() -> dict[str, tuple[int, int]]:
    """`código de nivel 3` → casilla `(fila, columna)` de «00 Datos Objeto».

    Se tiende sobre la plantilla **española**, que es de la que salieron los
    nombres del catálogo, y vale para todos los idiomas porque la rejilla ocupa
    las mismas casillas en todos.
    """
    libro = load_workbook(PLANTILLAS / FICHERO["es"])
    hoja = libro.worksheets[POS_OBJETO]
    # (capítulo, nombre normalizado) → casilla
    por_nombre: dict[tuple[str, str], tuple[int, int]] = {}
    for capitulo, fila in FILA_OBJETO.items():
        for i in range(MAX_OBJETOS):
            columna = COL_PRIMER_OBJETO + i
            etiqueta = _texto(hoja.cell(fila, columna))
            if etiqueta is not None:
                por_nombre.setdefault((capitulo, _normalizar(etiqueta)), (fila, columna))
    libro.close()

    casillas: dict[str, tuple[int, int]] = {}
    for codigo, (capitulo, nombre) in _nombres_del_catalogo().items():
        if capitulo not in FILA_OBJETO:
            continue  # los capítulos de soft costs no tienen lista de objetos
        casilla = por_nombre.get((capitulo, _normalizar(nombre)))
        if casilla is None:
            raise FaltaEnLaPlantilla(
                f"el catálogo trae {codigo} «{nombre}» en {capitulo}, y la plantilla "
                "española no lo tiene en esa lista"
            )
        casillas[codigo] = casilla
    return casillas


@lru_cache(maxsize=1)
def _rejilla_de_zonas() -> dict[tuple[int, int], str]:
    """`(fila, columna relativa)` → código de `zone`, leído del **español**.

    La plantilla no lleva códigos, solo texto, así que el puente entre sus
    casillas y `zone.code` se tiende una vez sobre la plantilla española —la
    que dio origen al catálogo sembrado— y a partir de ahí vale para todos los
    idiomas, porque la rejilla ocupa las mismas casillas en todos.
    """
    libro = load_workbook(PLANTILLAS / FICHERO["es"])
    activo = libro.worksheets[POS_ACTIVO]
    rejilla: dict[tuple[int, int], str] = {}
    for i in range(TIPOS):
        for fila in range(FILA_PRIMERA_ZONA, FILA_ULTIMA_ZONA + 1):
            etiqueta = _texto(activo.cell(fila, COL_PRIMER_TIPO + i))
            if etiqueta is None:
                continue
            sin_tildes = etiqueta.strip().lower().translate(str.maketrans("áéíóúü", "aeiouu"))
            normal = sin_tildes.replace(" ", "_")
            rejilla[(fila, i)] = CODIGO_DE_ZONA.get(normal, normal.upper())
    libro.close()
    return rejilla


#: `[SUP]` Etiqueta española normalizada → `zone.code`.
CODIGO_DE_ZONA = {
    "cuartos_tecnicos": "CUARTOS_TECNICOS",
    "aparcamiento": "APARCAMIENTO",
    "oficinas": "OFICINAS",
    "aseos": "ASEOS",
    "cubierta": "CUBIERTA",
    "zonas_exteriores": "ZONAS_EXTERIORES",
    "vestibulo_principal": "VESTIBULO_PRINCIPAL",
    "nucleo_escaleras": "NUCLEO_ESCALERAS",
    "general": "GENERAL",
    "vestibulo_de_planta": "VESTIBULO_PLANTA",
    "salas_de_personal": "SALAS_PERSONAL",
    "almacen": "ALMACEN",
    "vestuarios": "VESTUARIOS",
    "habitaciones": "HABITACIONES",
    "cocina": "COCINA",
    # `[REQ]` §5.2 · La plantilla escribe «Restaurante» en Hotel y Sanitario y
    # «Restaurantes» en Comercial. El catálogo lo unificó a propósito en una
    # sola zona: dos filas serían dos identificadores para el mismo concepto y
    # cualquier comparación de cartera daría dos líneas donde debe dar una. Las
    # dos casillas apuntan por eso al mismo código, y al exportar cada una
    # recupera **su** literal, que es el que espera su desplegable.
    "restaurante": "RESTAURANTE",
    "restaurantes": "RESTAURANTE",
    "gimnasio": "GIMNASIO",
    "piscina": "PISCINA",
    "zona_comercial": "ZONA_COMERCIAL",
    "salas_uso_sanitario": "SALAS_USO_SANITARIO",
}
