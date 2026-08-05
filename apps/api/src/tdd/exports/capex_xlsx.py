"""Hoja `CAPEX` del Excel exportado `[REQ]` P-31.

Consume **el mismo `CapexTableLayout`** que la tabla nativa del PPTX. Es lo que
garantiza que el Excel que el equipo adjunta en un correo y el PowerPoint que
va en ese mismo correo no tengan columnas distintas.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from tdd.reporting.capex_layout import TITULO_GRUPO, CapexTableLayout

VERDE = "A9C78C"
GRIS = "B0B0B0"
ORO = "9A8C4E"
GRIS_SECCION = "A6A6A6"
COLOR_PLAZO = {
    "corto": "F8CBCB",
    "medio": "FBE5A6",
    "largo": "C8E6C9",
    "mejoras": "BDD7EE",
    "otro": "E0E0E0",
}

_BORDE = Border(*[Side(style="thin", color="808080")] * 4)


def escribir_hoja(wb: Workbook, layout: CapexTableLayout) -> None:
    """Escribe la hoja `CAPEX` con el mismo layout que la tabla del informe."""
    ws = wb.active if wb.active and wb.active.max_row == 1 else wb.create_sheet()
    ws.title = "CAPEX"
    n = len(layout.columnas)

    # Fila 1: título del bloque, combinado
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1, value=layout.titulo)
    c.fill = PatternFill("solid", fgColor=VERDE)
    c.font = Font(bold=True, color="FFFFFF", name="Gotham Medium")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Filas 2-3: cabecera de dos niveles
    primera = next(i for i, col in enumerate(layout.columnas, 1) if col.grupo == "capex")
    ultima = max(i for i, col in enumerate(layout.columnas, 1) if col.grupo == "capex")
    ws.merge_cells(start_row=2, start_column=primera, end_row=2, end_column=ultima)
    cg = ws.cell(
        row=2,
        column=primera,
        value=TITULO_GRUPO["capex"][0 if layout.locale.startswith("es") else 1],
    )
    cg.fill = PatternFill("solid", fgColor=ORO)
    cg.font = Font(bold=True, color="FFFFFF", name="Gotham Medium")
    cg.alignment = Alignment(horizontal="center")

    for i, col in enumerate(layout.columnas, 1):
        if col.grupo is None:
            ws.merge_cells(start_row=2, start_column=i, end_row=3, end_column=i)
            celda = ws.cell(row=2, column=i, value=layout.titulo_columna(col))
            celda.fill = PatternFill("solid", fgColor=ORO if col.key == "riesgo" else GRIS)
            celda.font = Font(
                bold=True, name="Gotham Medium", color="FFFFFF" if col.key == "riesgo" else "000000"
            )
        else:
            celda = ws.cell(row=3, column=i, value=layout.titulo_columna(col))
            celda.fill = PatternFill("solid", fgColor=COLOR_PLAZO.get(col.key, GRIS))
            celda.font = Font(name="Gotham Light")
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = _BORDE
        # 1 pulgada ≈ 12 unidades de anchura de Excel
        ws.column_dimensions[get_column_letter(i)].width = round(col.ancho_in * 12, 1)

    # Cuerpo
    for f, fila in enumerate(layout.filas, start=4):
        resaltada = fila.tipo in ("seccion", "total")
        for i, col in enumerate(layout.columnas, 1):
            celda = ws.cell(row=f, column=i, value=fila.celdas.get(col.key, "") or None)
            celda.border = _BORDE
            celda.font = Font(
                bold=resaltada,
                color="FFFFFF" if resaltada else "000000",
                name="Gotham Medium" if resaltada else "Gotham Light",
            )
            if resaltada:
                celda.fill = PatternFill("solid", fgColor=GRIS_SECCION)
            celda.alignment = Alignment(
                horizontal=col.alineacion.value, vertical="top", wrap_text=not col.es_importe
            )

    ws.freeze_panes = "A4"


def generar_xlsx(layout: CapexTableLayout) -> bytes:
    """Devuelve el libro en memoria. La hoja `CAPEX` es la que se abre primero."""
    wb = Workbook()
    escribir_hoja(wb, layout)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
