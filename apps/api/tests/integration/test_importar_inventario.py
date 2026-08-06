"""Importación del inventario desde XLSX, punta a punta `[REQ]` §7 / P-15.

Lo que solo se ve contra PostgreSQL y con ficheros de verdad:

* que **previsualizar no escriba nada** —se comprueba contando filas antes y
  después, no leyendo el código—;
* que aplicar exija confirmación;
* que un equipo que ya está en la base **no se sobrescriba** por venir en una
  hoja, salvo que se pida.
"""

from __future__ import annotations

import uuid
from io import BytesIO
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import Engine, text

pytestmark = pytest.mark.db

RUTA = "/api/v1"

CABECERA = [
    "Activo",
    "Etiqueta",
    "Tipo de equipo",
    "Sistema técnico",
    "Fabricante",
    "Año de instalación",
    "Vida útil esperada",
    "Estado de conservación",
]


def libro(filas: list[list[Any]], cabecera: list[str] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(cabecera if cabecera is not None else CABECERA)
    for f in filas:
        ws.append(f)
    salida = BytesIO()
    wb.save(salida)
    return salida.getvalue()


def enviar(cliente: TestClient, cab: Any, proyecto: str, contenido: bytes, ruta: str, **datos: Any):
    return cliente.post(
        f"{RUTA}/projects/{proyecto}/equipment/{ruta}",
        headers=cab("consultor_a"),
        files={"file": ("inventario.xlsx", contenido, "application/vnd.ms-excel")},
        data={k: str(v).lower() for k, v in datos.items()},
    )


@pytest.fixture
def proyecto(datos_base: dict[str, uuid.UUID]) -> str:
    return str(datos_base["proyecto_a"])


@pytest.fixture
def nave(cliente: TestClient, cab: Any, proyecto: str, motor_admin: Engine) -> dict[str, str]:
    """Un activo nuevo por prueba.

    Devuelve el `id` además del nombre a propósito: el encargo es compartido
    entre pruebas y buscar «el equipo con etiqueta CL-01» en todo el proyecto
    encuentra el que dejó otra. Los listados de aquí filtran por activo.
    """
    with motor_admin.begin() as conn:
        tipologia = conn.execute(text("SELECT id FROM asset_typology LIMIT 1")).scalar_one()
    r = cliente.post(
        f"{RUTA}/projects/{proyecto}/assets",
        headers=cab("consultor_a"),
        json={
            "name": f"Nave {uuid.uuid4().hex[:6]}",
            "typology_id": str(tipologia),
            "asset_code": f"N-{uuid.uuid4().hex[:5]}",
        },
    )
    assert r.status_code == 201, r.text
    return {"id": r.json()["id"], "name": r.json()["name"]}


def inventario(cliente: TestClient, cab: Any, proyecto: str, activo: str) -> list[dict[str, Any]]:
    return cliente.get(
        f"{RUTA}/projects/{proyecto}/equipment?asset_id={activo}", headers=cab("consultor_a")
    ).json()


# ─────────────────────────────────────────────────────────────────────────────
#  La plantilla
# ─────────────────────────────────────────────────────────────────────────────


def test_la_plantilla_lleva_los_activos_del_encargo_y_los_sistemas(
    cliente: TestClient, cab: Any, proyecto: str, nave: dict[str, str]
) -> None:
    """Sin eso, quien la rellena escribe el nombre del edificio de memoria y la
    mitad de las filas fallan al importar por una tilde."""
    r = cliente.get(
        f"{RUTA}/projects/{proyecto}/equipment/import/plantilla.xlsx",
        headers=cab("consultor_a"),
    )
    assert r.status_code == 200, r.text
    assert "attachment" in r.headers["content-disposition"]

    wb = load_workbook(BytesIO(r.content))
    assert wb.worksheets[0]["A1"].value == "Activo"
    valores = "\n".join(
        str(c.value) for fila in wb["Valores admitidos"].iter_rows() for c in fila if c.value
    )
    assert nave["name"] in valores
    assert "Climatización" in valores
    # Y la lista de alias, para que nadie tenga que adivinar cómo llamar a una
    # columna cuando la hoja viene de fuera.
    assert "Cabeceras admitidas" in wb.sheetnames


def test_la_plantilla_se_puede_rellenar_e_importar_tal_cual(
    cliente: TestClient, cab: Any, proyecto: str, nave: dict[str, str]
) -> None:
    """La ida y la vuelta tienen que casar: una plantilla cuyos títulos no
    reconoce su propio importador es una trampa."""
    descargada = cliente.get(
        f"{RUTA}/projects/{proyecto}/equipment/import/plantilla.xlsx",
        headers=cab("consultor_a"),
    ).content
    wb = load_workbook(BytesIO(descargada))
    ws = wb.worksheets[0]
    ws.append([nave["name"], "CL-99", "Enfriadora", "Climatización"])
    salida = BytesIO()
    wb.save(salida)

    r = enviar(cliente, cab, proyecto, salida.getvalue(), "import/preview")
    assert r.status_code == 200, r.text
    assert r.json()["columnas_ignoradas"] == []
    assert r.json()["nuevas"] == 1
    # Las dos filas de ayuda de la plantilla se rechazan solas, sin que nadie
    # tenga que acordarse de borrarlas.
    assert r.json()["con_error"] == 2


# ─────────────────────────────────────────────────────────────────────────────
#  Previsualizar no escribe
# ─────────────────────────────────────────────────────────────────────────────


def test_previsualizar_no_guarda_nada(
    cliente: TestClient, cab: Any, proyecto: str, nave: dict[str, str]
) -> None:
    antes = len(inventario(cliente, cab, proyecto, nave["id"]))
    contenido = libro(
        [[nave["name"], "CL-01", "Enfriadora", "Climatización", "Ficticio", 2010, 20, "Bueno"]]
    )

    r = enviar(cliente, cab, proyecto, contenido, "import/preview")
    assert r.status_code == 200, r.text
    assert r.json()["nuevas"] == 1
    assert "Nada se ha guardado todavía" in r.json()["aviso"]

    assert len(inventario(cliente, cab, proyecto, nave["id"])) == antes


def test_la_previsualizacion_dice_fila_a_fila_que_va_a_pasar(
    cliente: TestClient, cab: Any, proyecto: str, nave: dict[str, str]
) -> None:
    contenido = libro(
        [
            [nave["name"], "CL-01", "Enfriadora", "Climatización"],
            ["Edificio Inexistente", "CL-02", "Bomba"],
            [nave["name"], "CL-01", "Otra enfriadora"],
        ]
    )
    datos = enviar(cliente, cab, proyecto, contenido, "import/preview").json()
    estados = {f["fila"]: f["estado"] for f in datos["filas"]}
    assert estados == {2: "NUEVA", 3: "ERROR", 4: "DUPLICADA_EN_FICHERO"}
    assert datos["nuevas"] == 1
    assert datos["con_error"] == 2


def test_un_fichero_que_no_es_un_xlsx_se_rechaza_diciendo_que_hacer(
    cliente: TestClient, cab: Any, proyecto: str
) -> None:
    r = enviar(cliente, cab, proyecto, b"esto no es un libro", "import/preview")
    assert r.status_code == 422
    assert ".xlsx" in r.json()["detail"]


# ─────────────────────────────────────────────────────────────────────────────
#  Aplicar
# ─────────────────────────────────────────────────────────────────────────────


def test_importar_sin_confirmar_no_hace_nada(
    cliente: TestClient, cab: Any, proyecto: str, nave: dict[str, str]
) -> None:
    antes = len(inventario(cliente, cab, proyecto, nave["id"]))
    contenido = libro([[nave["name"], "CL-01", "Enfriadora"]])
    r = enviar(cliente, cab, proyecto, contenido, "import")
    assert r.status_code == 422
    assert "previsualización" in r.json()["detail"]
    assert len(inventario(cliente, cab, proyecto, nave["id"])) == antes


def test_importar_confirmando_crea_los_equipos_con_su_vida_calculada(
    cliente: TestClient, cab: Any, proyecto: str, nave: dict[str, str]
) -> None:
    contenido = libro(
        [
            [nave["name"], "CL-01", "Enfriadora", "Climatización", "Ficticio", 1995, 20, "Bueno"],
            [nave["name"], "AS-01", "Ascensor", "Ascensores", "", "", "", ""],
        ]
    )
    r = enviar(cliente, cab, proyecto, contenido, "import", confirmar=True)
    assert r.status_code == 200, r.text
    assert r.json()["creados"] == 2

    equipos = {e["tag"]: e for e in inventario(cliente, cab, proyecto, nave["id"])}
    assert equipos["CL-01"]["technical_system_name"] == "Climatización"
    assert equipos["CL-01"]["end_of_life_year"] == 2015
    assert equipos["CL-01"]["vencido"] is True
    # Y el que vino sin años no finge un cálculo.
    assert equipos["AS-01"]["remaining_life_years"] is None


def test_las_filas_con_error_no_impiden_que_entren_las_buenas(
    cliente: TestClient, cab: Any, proyecto: str, nave: dict[str, str]
) -> None:
    """Rechazar la hoja entera por una fila mal escrita obligaría a repetir el
    trabajo de las otras doscientas."""
    contenido = libro(
        [[nave["name"], "OK-01", "Enfriadora"], ["Edificio Inexistente", "MAL-01", "Bomba"]]
    )
    r = enviar(cliente, cab, proyecto, contenido, "import", confirmar=True)
    assert r.json() == {**r.json(), "creados": 1, "omitidos": 1}

    etiquetas = {e["tag"] for e in inventario(cliente, cab, proyecto, nave["id"])}
    assert "OK-01" in etiquetas
    assert "MAL-01" not in etiquetas


# ─────────────────────────────────────────────────────────────────────────────
#  Nada se sobrescribe solo
# ─────────────────────────────────────────────────────────────────────────────


def test_reimportar_no_duplica_ni_pisa_lo_que_ya_hay(
    cliente: TestClient, cab: Any, proyecto: str, nave: dict[str, str]
) -> None:
    """`[REQ]` La ficha que hay en la base la escribió alguien en una visita a
    la que no se vuelve. Importar dos veces la misma hoja —que pasa— no puede
    borrar lo que se corrigió a mano entre medias."""
    enviar(
        cliente,
        cab,
        proyecto,
        libro([[nave["name"], "CL-01", "Enfriadora"]]),
        "import",
        confirmar=True,
    )

    equipo = next(e for e in inventario(cliente, cab, proyecto, nave["id"]) if e["tag"] == "CL-01")
    cliente.patch(
        f"{RUTA}/equipment/{equipo['id']}",
        headers=cab("consultor_a"),
        json={"manufacturer": "Corregido a mano tras la visita"},
    )

    segunda = enviar(
        cliente,
        cab,
        proyecto,
        libro([[nave["name"], "CL-01", "Enfriadora", "", "De la hoja"]]),
        "import",
        confirmar=True,
    )
    assert segunda.json()["creados"] == 0
    assert segunda.json()["actualizados"] == 0
    assert segunda.json()["omitidos"] == 1

    despues = cliente.get(f"{RUTA}/equipment/{equipo['id']}", headers=cab("consultor_a")).json()
    assert despues["manufacturer"] == "Corregido a mano tras la visita"


def test_actualizar_los_existentes_es_una_decision_explicita(
    cliente: TestClient, cab: Any, proyecto: str, nave: dict[str, str]
) -> None:
    enviar(
        cliente,
        cab,
        proyecto,
        libro([[nave["name"], "CL-01", "Enfriadora"]]),
        "import",
        confirmar=True,
    )
    r = enviar(
        cliente,
        cab,
        proyecto,
        libro([[nave["name"], "CL-01", "Enfriadora", "", "De la hoja"]]),
        "import",
        confirmar=True,
        actualizar_existentes=True,
    )
    assert r.json()["actualizados"] == 1

    equipo = next(e for e in inventario(cliente, cab, proyecto, nave["id"]) if e["tag"] == "CL-01")
    assert equipo["manufacturer"] == "De la hoja"


def test_la_importacion_queda_auditada(
    cliente: TestClient, cab: Any, proyecto: str, nave: dict[str, str], motor_admin: Engine
) -> None:
    """Una importación mueve muchas filas de golpe: es justo lo que alguien
    querrá reconstruir dentro de seis meses."""
    enviar(
        cliente,
        cab,
        proyecto,
        libro([[nave["name"], "AUD-01", "Enfriadora"]]),
        "import",
        confirmar=True,
    )
    with motor_admin.begin() as conn:
        fila = conn.execute(
            text(
                "SELECT after_data FROM audit_log WHERE action = 'EQUIPMENT_IMPORTED' "
                "AND entity_id = :p ORDER BY occurred_at DESC LIMIT 1"
            ),
            {"p": proyecto},
        ).first()
    assert fila is not None
    assert fila[0]["creados"] >= 1


def test_otra_organizacion_no_importa_en_este_encargo(
    cliente: TestClient, cab: Any, proyecto: str, nave: dict[str, str]
) -> None:
    r = cliente.post(
        f"{RUTA}/projects/{proyecto}/equipment/import/preview",
        headers=cab("admin_b"),
        files={
            "file": (
                "i.xlsx",
                libro([[nave["name"], "X", "Enfriadora"]]),
                "application/vnd.ms-excel",
            )
        },
    )
    # La RLS oculta los activos del encargo ajeno, así que ninguna fila casa.
    assert r.json()["nuevas"] == 0
