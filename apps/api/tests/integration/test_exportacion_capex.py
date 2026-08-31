"""`[REQ]` P-31 · El botón de exportar el CAPEX a XLSX.

> *«añade en la aplicación un botón de exportar el CAPEX a xlsx para que el
> equipo pueda adjuntar el fichero en el posterior envío que hagan fuera de la
> plataforma»*

El generador del libro estaba escrito y probado desde el principio, pero **no
había ninguna ruta que lo sirviera**: el botón no habría tenido a dónde llamar.
Se detectó recorriendo la aplicación en marcha, y estas pruebas son lo que lo
habría detectado antes.

Lo que se comprueba no es solo que devuelva 200: que el fichero sea un libro
de verdad, que **incluya las líneas en borrador** —el equipo comparte el CAPEX
mientras lo está construyendo— y que **excluya las descartadas**.
"""

from __future__ import annotations

import io
import uuid
import zipfile
from decimal import Decimal
from typing import Any

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import Engine, text

pytestmark = pytest.mark.db

RUTA = "/api/v1"
EXPORTACION = f"{RUTA}/projects/{{}}/capex/export.xlsx"
TIPO_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(scope="module")
def catalogo(motor_admin: Engine) -> dict[str, str]:
    with motor_admin.begin() as conn:
        tipologia = conn.execute(
            text("SELECT id FROM asset_typology ORDER BY sort_order LIMIT 1")
        ).scalar_one()
        zona = conn.execute(
            text(
                "SELECT z.id FROM zone z JOIN zone_typology zt ON zt.zone_id = z.id "
                "WHERE zt.typology_id = :t ORDER BY z.sort_order LIMIT 1"
            ),
            {"t": tipologia},
        ).scalar_one()
        codigo = conn.execute(
            text("SELECT id FROM capex_code WHERE level = 3 ORDER BY code LIMIT 1")
        ).scalar_one()
    return {"tipologia": str(tipologia), "zona": str(zona), "codigo": str(codigo)}


@pytest.fixture
def proyecto(motor_admin: Engine, datos_base: dict[str, uuid.UUID]) -> str:
    with motor_admin.begin() as conn:
        return str(
            conn.execute(
                text(
                    "INSERT INTO project (organization_id, client_id, internal_code, name) "
                    "VALUES (:o, :c, :cod, 'Encargo exportable') RETURNING id"
                ),
                {
                    "o": str(datos_base["org_a"]),
                    "c": str(datos_base["cliente_a"]),
                    "cod": f"EXP-{uuid.uuid4().hex[:6]}",
                },
            ).scalar_one()
        )


@pytest.fixture
def activo(cliente: TestClient, cab: Any, proyecto: str, catalogo: dict[str, str]) -> str:
    return str(
        cliente.post(
            f"{RUTA}/projects/{proyecto}/assets",
            headers=cab("consultor_a"),
            json={"name": "Nave A", "typology_id": catalogo["tipologia"]},
        ).json()["id"]
    )


def crear_hallazgo(
    cliente: TestClient,
    cab: Any,
    proyecto: str,
    catalogo: dict[str, str],
    activo: str,
    *,
    titulo: str,
    importe: str,
) -> str:
    r = cliente.post(
        f"{RUTA}/projects/{proyecto}/findings",
        headers=cab("consultor_a"),
        json={
            "asset_id": activo,
            "capex_code_id": catalogo["codigo"],
            "zone_id": catalogo["zona"],
            "title": titulo,
            "description": "Anotado en visita.",
            "capex_lines": [{"time_horizon_code": "CORTO", "amount": importe}],
        },
    )
    assert r.status_code == 201, r.text
    return str(r.json()["id"])


def hojas(contenido: bytes) -> list[str]:
    """Los nombres de hoja, leyendo el XLSX como el ZIP que es.

    Se abre el fichero de verdad en vez de fiarse del `Content-Type`: una
    cabecera correcta sobre un cuerpo corrupto es exactamente el fallo que un
    adjunto de correo no perdona.
    """
    import re
    from xml.sax.saxutils import unescape

    with zipfile.ZipFile(io.BytesIO(contenido)) as z:
        libro = z.read("xl/workbook.xml").decode("utf-8")
    # Solo los `<sheet name=...>`: `name=` también lo llevan los nombres
    # definidos, y la plantilla del cliente trae veintitantos.
    return [unescape(n) for n in re.findall(r'<sheet name="([^"]+)"', libro)]


def test_el_libro_se_descarga_y_se_abre(
    cliente: TestClient, cab: Any, proyecto: str, catalogo: dict[str, str], activo: str
) -> None:
    crear_hallazgo(
        cliente, cab, proyecto, catalogo, activo, titulo="Cubierta agotada", importe="36125.00"
    )

    r = cliente.get(EXPORTACION.format(proyecto), headers=cab("consultor_a"))

    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == TIPO_XLSX
    assert r.content[:2] == b"PK", "no es un ZIP: el libro está corrupto"
    # Sale la plantilla del cliente entera, no una hoja construida a mano: sus
    # siete hojas con los gráficos, las dinámicas y las listas.
    assert hojas(r.content) == [
        "00 Datos Categorías",
        "00 Datos Objeto",
        "00 Datos Activo",
        "Leyenda",
        "CapEx",
        "Resumen CapEx",
        "Gráficas & Desglose Hard Costs",
    ]


def test_el_nombre_del_fichero_lleva_el_codigo_del_encargo(
    cliente: TestClient,
    cab: Any,
    proyecto: str,
    catalogo: dict[str, str],
    activo: str,
    motor_admin: Engine,
) -> None:
    """El equipo adjunta varios en el mismo correo: `export.xlsx` repetido no
    sirve para nada."""
    crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo="Fisuras", importe="1200.00")
    with motor_admin.begin() as conn:
        codigo = conn.execute(
            text("SELECT internal_code FROM project WHERE id = :p"), {"p": proyecto}
        ).scalar_one()

    r = cliente.get(EXPORTACION.format(proyecto), headers=cab("consultor_a"))
    assert codigo in r.headers["content-disposition"]


def test_un_encargo_sin_capex_no_devuelve_un_libro_vacio(
    cliente: TestClient, cab: Any, proyecto: str
) -> None:
    """`[REQ]` Un Excel con la cabecera y ninguna fila se adjunta a un correo
    sin que nadie note que no lleva nada dentro."""
    r = cliente.get(EXPORTACION.format(proyecto), headers=cab("consultor_a"))
    assert r.status_code == 409
    assert "actuación" in r.json()["detail"]


def test_incluye_los_hallazgos_en_borrador(
    cliente: TestClient, cab: Any, proyecto: str, catalogo: dict[str, str], activo: str
) -> None:
    """`[REQ]` El informe deja fuera los borradores; **esta exportación no**.

    Es un fichero de trabajo: si se dejara fuera lo que aún está en borrador, el
    total del Excel no cuadraría con el de la pantalla desde la que se pulsó el
    botón, y nadie sabría cuál de los dos creer.
    """
    crear_hallazgo(
        cliente, cab, proyecto, catalogo, activo, titulo="Recién anotado", importe="7500.00"
    )
    # Sin ninguna transición: el hallazgo sigue en BORRADOR.
    r = cliente.get(EXPORTACION.format(proyecto), headers=cab("consultor_a"))
    assert r.status_code == 200, r.text


def test_lo_descartado_se_queda_fuera(
    cliente: TestClient,
    cab: Any,
    proyecto: str,
    catalogo: dict[str, str],
    activo: str,
    motor_admin: Engine,
) -> None:
    """Descartar una actuación es decir que no se hace. Sumarla al total del
    fichero que se envía la resucitaría por la puerta de atrás."""
    descartado = crear_hallazgo(
        cliente, cab, proyecto, catalogo, activo, titulo="Descartado", importe="99999.00"
    )
    with motor_admin.begin() as conn:
        conn.execute(
            text("UPDATE finding SET status = 'DESCARTADO' WHERE id = :i"), {"i": descartado}
        )

    r = cliente.get(EXPORTACION.format(proyecto), headers=cab("consultor_a"))
    assert r.status_code == 409, "sin más hallazgos, el único que quedaba estaba descartado"


def test_otra_organizacion_no_exporta_el_capex_ajeno(
    cliente: TestClient, cab: Any, proyecto: str, catalogo: dict[str, str], activo: str
) -> None:
    crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo="Privado", importe="500.00")
    r = cliente.get(EXPORTACION.format(proyecto), headers=cab("admin_b"))
    assert r.status_code in (404, 409), r.text
    assert r.status_code != 200


def test_el_lector_tambien_puede_exportar(
    cliente: TestClient, cab: Any, proyecto: str, catalogo: dict[str, str], activo: str
) -> None:
    """`[SUP]` Exportar es leer. Un `LECTOR` que ve la tabla en pantalla puede
    llevársela en Excel: negárselo no protegería nada."""
    crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo="Legible", importe="800.00")
    r = cliente.get(EXPORTACION.format(proyecto), headers=cab("lector_a"))
    assert r.status_code == 200, r.text


# ─────────────────────────────────────────────────────────────────────────────
#  La plantilla del cliente, rellenada
# ─────────────────────────────────────────────────────────────────────────────


def _capex(contenido: bytes):
    from openpyxl import load_workbook

    return load_workbook(io.BytesIO(contenido))["CapEx"]


def test_la_actuacion_llega_a_su_bloque_de_la_plantilla(
    cliente: TestClient, cab: Any, proyecto: str, catalogo: dict[str, str], activo: str
) -> None:
    """No basta con que el fichero se abra: hay que ver el dato dentro, y en la
    fila que le toca por su capítulo."""
    crear_hallazgo(
        cliente, cab, proyecto, catalogo, activo, titulo="Cubierta agotada", importe="36125.00"
    )
    r = cliente.get(EXPORTACION.format(proyecto), headers=cab("consultor_a"))
    hoja = _capex(r.content)

    descripciones = [hoja[f"G{fila}"].value for fila in range(14, 255) if hoja[f"G{fila}"].value]
    assert "Cubierta agotada" in descripciones


def test_el_idioma_elige_la_plantilla(
    cliente: TestClient, cab: Any, proyecto: str, catalogo: dict[str, str], activo: str
) -> None:
    """`[REQ]` Al extraer el informe se elige español o inglés, y el Excel tiene
    que salir en el mismo idioma que el PowerPoint que lo acompaña."""
    crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo="Roof", importe="1000.00")

    es = cliente.get(EXPORTACION.format(proyecto), headers=cab("consultor_a"))
    en = cliente.get(f"{EXPORTACION.format(proyecto)}?idioma=en", headers=cab("consultor_a"))
    assert (es.status_code, en.status_code) == (200, 200)

    from openpyxl import load_workbook

    assert load_workbook(io.BytesIO(es.content)).sheetnames[0] == "00 Datos Categorías"
    assert load_workbook(io.BytesIO(en.content)).sheetnames[0] == "00 Category Data"


def test_un_idioma_sin_plantilla_se_rechaza_con_422(
    cliente: TestClient, cab: Any, proyecto: str
) -> None:
    """Y no con un 500 al no encontrar el fichero: el error es de quien llama."""
    r = cliente.get(f"{EXPORTACION.format(proyecto)}?idioma=fr", headers=cab("consultor_a"))
    assert r.status_code == 422
    assert "es" in r.text and "en" in r.text


def test_mas_actuaciones_de_las_que_caben_dan_409_y_dicen_cual(
    cliente: TestClient, cab: Any, proyecto: str, catalogo: dict[str, str], activo: str
) -> None:
    """`[LIM]` Cada capítulo admite diez filas. La undécima no cabe, y lo que no
    puede pasar es que se pierda en silencio: se corta la descarga diciendo qué
    capítulo se pasa y por cuánto."""
    for i in range(11):
        crear_hallazgo(
            cliente, cab, proyecto, catalogo, activo, titulo=f"Actuación {i}", importe="100.00"
        )
    r = cliente.get(EXPORTACION.format(proyecto), headers=cab("consultor_a"))
    assert r.status_code == 409
    assert "11 de 10" in r.json()["detail"]


# ─────────────────────────────────────────────────────────────────────────────
#  Carteras: el CAPEX separado por activo
# ─────────────────────────────────────────────────────────────────────────────
#
# La plantilla del cliente describe UN edificio. Metidos tres activos en un solo
# libro, la cabecera describe al primero y los otros dos salen sin identificar.
# Estas pruebas cubren la salida: un libro por activo.

CARTERA_ZIP = f"{RUTA}/projects/{{}}/capex/export.zip"
POR_ACTIVO = f"{RUTA}/projects/{{}}/capex/summary/by-asset"


@pytest.fixture
def segundo_activo(cliente: TestClient, cab: Any, proyecto: str, catalogo: dict[str, str]) -> str:
    return str(
        cliente.post(
            f"{RUTA}/projects/{proyecto}/assets",
            headers=cab("consultor_a"),
            json={
                "name": "Nave B",
                "asset_code": "NB-02",
                "typology_id": catalogo["tipologia"],
            },
        ).json()["id"]
    )


def libros_del_zip(contenido: bytes) -> dict[str, bytes]:
    with zipfile.ZipFile(io.BytesIO(contenido)) as z:
        return {n: z.read(n) for n in z.namelist()}


def test_la_cartera_se_descarga_con_un_libro_por_activo(
    cliente: TestClient,
    cab: Any,
    proyecto: str,
    catalogo: dict[str, str],
    activo: str,
    segundo_activo: str,
) -> None:
    """`[REQ]` Lo que pedía el cliente: el CAPEX **separado por activo**."""
    crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo="Cubierta A", importe="100")
    crear_hallazgo(
        cliente, cab, proyecto, catalogo, segundo_activo, titulo="Cubierta B", importe="200"
    )

    r = cliente.get(CARTERA_ZIP.format(proyecto), headers=cab("consultor_a"))

    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == "application/zip"
    dentro = libros_del_zip(r.content)
    libros = sorted(n for n in dentro if n.endswith(".xlsx"))
    assert len(libros) == 2, dentro.keys()
    # Cada uno es la plantilla del cliente entera, no una hoja recortada.
    for nombre in libros:
        assert hojas(dentro[nombre])[0] == "00 Datos Categorías"


def test_cada_libro_lleva_solo_las_actuaciones_de_su_activo(
    cliente: TestClient,
    cab: Any,
    proyecto: str,
    catalogo: dict[str, str],
    activo: str,
    segundo_activo: str,
) -> None:
    """Es la propiedad que hace útil la separación.

    Si las actuaciones de los dos activos aparecieran en los dos libros, el
    total de cada edificio estaría inflado y nadie lo vería hasta cuadrar la
    suma a mano contra el informe.
    """
    crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo="Solo en A", importe="100")
    crear_hallazgo(
        cliente, cab, proyecto, catalogo, segundo_activo, titulo="Solo en B", importe="200"
    )

    dentro = libros_del_zip(
        cliente.get(CARTERA_ZIP.format(proyecto), headers=cab("consultor_a")).content
    )
    # Solo las descripciones que escribe la aplicación: la columna G lleva
    # también las filas de soft costs que trae la propia plantilla.
    nuestras = {"Solo en A", "Solo en B"}
    textos = [
        {v for fila in range(14, 255) if (v := _capex(contenido)[f"G{fila}"].value) in nuestras}
        for nombre, contenido in dentro.items()
        if nombre.endswith(".xlsx")
    ]

    assert len(textos) == 2
    for descripciones in textos:
        assert len(descripciones) == 1, f"un libro lleva actuaciones ajenas: {descripciones}"
    assert textos[0] | textos[1] == nuestras, "alguna actuación se ha perdido"


def test_el_libro_de_cada_activo_lleva_su_nombre_en_la_cabecera(
    cliente: TestClient,
    cab: Any,
    proyecto: str,
    catalogo: dict[str, str],
    activo: str,
    segundo_activo: str,
) -> None:
    """La celda se llama «Nombre del proyecto» y en una cartera no basta.

    La hoja `CapEx` y las gráficas la referencian por fórmula, así que el
    nombre se propaga solo al resto del libro.
    """
    from openpyxl import load_workbook

    crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo="A", importe="100")
    crear_hallazgo(cliente, cab, proyecto, catalogo, segundo_activo, titulo="B", importe="200")

    dentro = libros_del_zip(
        cliente.get(CARTERA_ZIP.format(proyecto), headers=cab("consultor_a")).content
    )
    nombres = {
        load_workbook(io.BytesIO(c))["00 Datos Activo"]["C5"].value
        for n, c in dentro.items()
        if n.endswith(".xlsx")
    }

    assert any(n and n.endswith("Nave A") for n in nombres), nombres
    assert any(n and n.endswith("Nave B") for n in nombres), nombres


def test_el_zip_declara_los_activos_que_se_quedaron_sin_libro(
    cliente: TestClient,
    cab: Any,
    proyecto: str,
    catalogo: dict[str, str],
    activo: str,
    segundo_activo: str,
) -> None:
    """`[REQ]` Un activo sin actuaciones no lleva libro, **y se dice**.

    Omitirlo en silencio haría que un edificio sin visitar y otro visitado sin
    hallazgos se vieran igual desde fuera del ZIP.
    """
    crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo="Solo A", importe="100")

    dentro = libros_del_zip(
        cliente.get(CARTERA_ZIP.format(proyecto), headers=cab("consultor_a")).content
    )

    assert len([n for n in dentro if n.endswith(".xlsx")]) == 1
    leeme = dentro["LEEME.txt"].decode("utf-8")
    assert "Nave B" in leeme
    assert "SIN libro" in leeme


def test_una_cartera_sin_ninguna_actuacion_no_devuelve_un_zip_vacio(
    cliente: TestClient, cab: Any, proyecto: str, activo: str
) -> None:
    r = cliente.get(CARTERA_ZIP.format(proyecto), headers=cab("consultor_a"))
    assert r.status_code == 409
    assert "actuación" in r.json()["detail"]


def test_la_cabida_de_la_plantilla_se_cuenta_por_activo(
    cliente: TestClient,
    cab: Any,
    proyecto: str,
    catalogo: dict[str, str],
    activo: str,
    segundo_activo: str,
) -> None:
    """`[REQ]` Diez actuaciones por capítulo **y por activo**.

    Doce repartidas entre dos naves revientan el libro conjunto y caben
    separadas. Es la otra mitad de por qué una cartera necesita separarse: no
    es solo la cabecera, es que si no, no cabe.
    """
    for i in range(6):
        crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo=f"A{i}", importe="100")
        crear_hallazgo(
            cliente, cab, proyecto, catalogo, segundo_activo, titulo=f"B{i}", importe="100"
        )

    junto = cliente.get(EXPORTACION.format(proyecto), headers=cab("consultor_a"))
    separado = cliente.get(CARTERA_ZIP.format(proyecto), headers=cab("consultor_a"))

    assert junto.status_code == 409, "doce en un capítulo no caben en un libro"
    assert separado.status_code == 200, separado.text
    assert len([n for n in libros_del_zip(separado.content) if n.endswith(".xlsx")]) == 2


def test_se_puede_descargar_el_libro_de_un_solo_activo(
    cliente: TestClient,
    cab: Any,
    proyecto: str,
    catalogo: dict[str, str],
    activo: str,
    segundo_activo: str,
) -> None:
    """El caso cotidiano: el consultor manda el CAPEX de una nave, no de todas."""
    crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo="Solo en A", importe="100")
    crear_hallazgo(
        cliente, cab, proyecto, catalogo, segundo_activo, titulo="Solo en B", importe="200"
    )

    r = cliente.get(
        f"{EXPORTACION.format(proyecto)}?asset_id={segundo_activo}", headers=cab("consultor_a")
    )

    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == TIPO_XLSX
    # `asset_code` en el nombre del fichero: el equipo adjunta varios a la vez.
    assert "NB-02" in r.headers["content-disposition"]
    hoja = _capex(r.content)
    descripciones = [hoja[f"G{f}"].value for f in range(14, 255) if hoja[f"G{f}"].value]
    assert "Solo en B" in descripciones
    assert "Solo en A" not in descripciones, "el libro de una nave lleva lo de la otra"


def test_un_activo_de_otro_encargo_da_404(
    cliente: TestClient, cab: Any, proyecto: str, catalogo: dict[str, str], activo: str
) -> None:
    """Y no un libro vacío ni el del encargo entero: el parámetro está mal."""
    crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo="A", importe="100")

    r = cliente.get(
        f"{EXPORTACION.format(proyecto)}?asset_id={uuid.uuid4()}", headers=cab("consultor_a")
    )
    assert r.status_code == 404


def test_un_activo_sin_actuaciones_lo_dice_con_su_nombre(
    cliente: TestClient,
    cab: Any,
    proyecto: str,
    catalogo: dict[str, str],
    activo: str,
    segundo_activo: str,
) -> None:
    """El mensaje nombra el activo, no «el encargo»: quien pulsa está en su ficha."""
    crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo="Solo A", importe="100")

    r = cliente.get(
        f"{EXPORTACION.format(proyecto)}?asset_id={segundo_activo}", headers=cab("consultor_a")
    )
    assert r.status_code == 409
    assert "Nave B" in r.json()["detail"]


def test_el_resumen_por_activo_suma_lo_de_cada_uno(
    cliente: TestClient,
    cab: Any,
    proyecto: str,
    catalogo: dict[str, str],
    activo: str,
    segundo_activo: str,
) -> None:
    """`[REQ]` El número que entra en la negociación de cada edificio."""
    crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo="A1", importe="1000.00")
    crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo="A2", importe="500.00")
    crear_hallazgo(cliente, cab, proyecto, catalogo, segundo_activo, titulo="B1", importe="2000.00")

    r = cliente.get(POR_ACTIVO.format(proyecto), headers=cab("consultor_a"))

    assert r.status_code == 200, r.text
    por_nombre = {fila["asset_name"]: fila for fila in r.json()}
    assert Decimal(por_nombre["Nave A"]["amount"]) == Decimal("1500.00")
    assert Decimal(por_nombre["Nave B"]["amount"]) == Decimal("2000.00")
    assert por_nombre["Nave A"]["findings"] == 2


def test_un_activo_sin_hallazgos_sale_en_el_resumen_con_ceros(
    cliente: TestClient,
    cab: Any,
    proyecto: str,
    catalogo: dict[str, str],
    activo: str,
    segundo_activo: str,
) -> None:
    """No es lo mismo un edificio sin visitar que uno visitado y sin hallazgos.

    Si desapareciera de la tabla, desde la pantalla se verían igual.
    """
    crear_hallazgo(cliente, cab, proyecto, catalogo, activo, titulo="A1", importe="1000.00")

    filas = cliente.get(POR_ACTIVO.format(proyecto), headers=cab("consultor_a")).json()

    vacio = next(f for f in filas if f["asset_name"] == "Nave B")
    assert (vacio["findings"], Decimal(vacio["amount"])) == (0, Decimal("0"))


def test_otra_organizacion_no_ve_el_resumen_por_activo(
    cliente: TestClient, cab: Any, proyecto: str, activo: str
) -> None:
    filas = cliente.get(POR_ACTIVO.format(proyecto), headers=cab("admin_b")).json()
    assert filas == []
