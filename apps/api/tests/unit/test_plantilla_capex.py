"""La plantilla CAPEX del cliente se rellena sin destruirla `[REQ]` P-31.

La prueba que sostiene el módulo es `test_no_se_pierde_ninguna_parte`. Volver a
escribir el libro con `openpyxl` destruye 33 de sus 87 partes —gráficos,
logotipos, segmentaciones, comentarios— y el resultado se abre igual: es la
peor forma posible de romperlo, porque nadie lo nota hasta que el cliente
pregunta dónde están sus gráficas.
"""

from __future__ import annotations

import zipfile
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from tdd.exports.plantilla_capex import (
    FICHERO,
    GEOMETRIA,
    PLANTILLAS,
    POR_CODIGO,
    Actuacion,
    CeldaInexistente,
    Encargo,
    NoCabe,
    comprobar_cabida,
    generar,
)
from tdd.exports.vocabulario_capex import leer

IDIOMAS = sorted(FICHERO)


@pytest.fixture
def encargo() -> Encargo:
    return Encargo(
        nombre="Encargo de prueba",
        direccion="Sin dirección real",
        fecha="2026-07-28",
        ano_construccion=1998,
        superficie_total=Decimal("18400"),
        tipo_edificio="INDUSTRIAL",
    )


def _actuacion(categoria: str = "HC.H08", **extra) -> Actuacion:
    base = {
        "objeto": "Producción de climatización",
        "zona": "Cuartos Técnicos",
        "descripcion": "Enfriadora fuera de reglamento",
        "riesgo": "04 Extremo",
        "importes": {"CORTO": Decimal("271700")},
    }
    return Actuacion(categoria=categoria, **{**base, **extra})


# ─────────────────────────────────────────────────────────────────────────────
#  Fidelidad del fichero
# ─────────────────────────────────────────────────────────────────────────────


@pytest.mark.parametrize("idioma", IDIOMAS)
def test_no_se_pierde_ninguna_parte(encargo: Encargo, idioma: str) -> None:
    """Ni una. La plantilla lleva gráficos, tablas dinámicas, segmentaciones,
    dos logotipos y formato condicional avanzado; reconstruir el libro los
    tiraría y devolvería al cliente su propia hoja rota."""
    original = zipfile.ZipFile(PLANTILLAS / FICHERO[idioma]).namelist()
    salida = zipfile.ZipFile(BytesIO(generar(encargo, [_actuacion()], idioma=idioma))).namelist()
    assert set(original) == set(salida)
    # Y en el mismo orden, que es lo que hace el fichero reproducible.
    assert original == salida


@pytest.mark.parametrize("idioma", IDIOMAS)
def test_sobreviven_las_extensiones_que_openpyxl_tira(encargo: Encargo, idioma: str) -> None:
    """Las validaciones y el formato condicional modernos viven en un `extLst`
    que `openpyxl` no entiende y elimina al guardar. Sin ellos los desplegables
    en cascada dejan de funcionar, y eso no se ve hasta que alguien teclea."""
    datos = generar(encargo, [_actuacion()], idioma=idioma)
    with zipfile.ZipFile(BytesIO(datos)) as z:
        hoja = next(n for n in z.namelist() if n.endswith("sheet5.xml"))
        crudo = z.read(hoja).decode("utf-8")
    assert "x14:dataValidation" in crudo
    assert "x14:conditionalFormatting" in crudo


@pytest.mark.parametrize("idioma", IDIOMAS)
def test_sale_un_xlsx_y_no_una_plantilla(encargo: Encargo, idioma: str) -> None:
    """El origen es `.xltm`. Se comprobó que **no lleva `vbaProject.bin`**, así
    que convertirlo a libro normal no pierde nada y evita que el correo del
    cliente lo bloquee por venir marcado como macros."""
    with zipfile.ZipFile(BytesIO(generar(encargo, [_actuacion()], idioma=idioma))) as z:
        tipos = z.read("[Content_Types].xml").decode("utf-8")
        assert "spreadsheetml.sheet.main+xml" in tipos
        assert "template.macroEnabled" not in tipos
        assert not [n for n in z.namelist() if "vbaProject" in n]


@pytest.mark.parametrize("idioma", IDIOMAS)
def test_excel_recalculara_al_abrir(encargo: Encargo, idioma: str) -> None:
    """Los subtotales de la plantilla traen cacheado el cero del libro en
    blanco. Sin forzar el recálculo, la hoja se abriría enseñando ceros con las
    líneas rellenas justo encima."""
    with zipfile.ZipFile(BytesIO(generar(encargo, [_actuacion()], idioma=idioma))) as z:
        assert 'fullCalcOnLoad="1"' in z.read("xl/workbook.xml").decode("utf-8")


# ─────────────────────────────────────────────────────────────────────────────
#  Dónde cae cada cosa
# ─────────────────────────────────────────────────────────────────────────────


def test_cada_actuacion_cae_en_el_bloque_de_su_categoria(encargo: Encargo) -> None:
    datos = generar(
        encargo,
        [
            _actuacion("HC.H02", objeto="Cubierta", descripcion="Lámina al final de su vida"),
            _actuacion("HC.H09", objeto="CGBT", descripcion="Cuadro sin diferencial"),
            _actuacion("MA.General", objeto="Ruido", descripcion="Nivel sonoro en fachada"),
        ],
    )
    hoja = load_workbook(BytesIO(datos))["CapEx"]
    assert hoja[f"G{POR_CODIGO['HC.H02'].primera}"].value == "Lámina al final de su vida"
    assert hoja[f"G{POR_CODIGO['HC.H09'].primera}"].value == "Cuadro sin diferencial"
    assert hoja[f"G{POR_CODIGO['MA.General'].primera}"].value == "Nivel sonoro en fachada"


def test_dos_actuaciones_de_la_misma_categoria_van_en_filas_seguidas(encargo: Encargo) -> None:
    datos = generar(
        encargo,
        [_actuacion("HC.H02", descripcion="Primera"), _actuacion("HC.H02", descripcion="Segunda")],
    )
    hoja = load_workbook(BytesIO(datos))["CapEx"]
    bloque = POR_CODIGO["HC.H02"]
    assert hoja[f"G{bloque.primera}"].value == "Primera"
    assert hoja[f"G{bloque.primera + 1}"].value == "Segunda"


def test_una_actuacion_recurrente_ocupa_una_columna_por_plazo(encargo: Encargo) -> None:
    """`[REQ]` P-44 · Una actuación con dos plazos lleva importe en las dos
    columnas de la misma fila, no dos filas."""
    datos = generar(
        encargo,
        [
            _actuacion(
                "HC.H02",
                descripcion="Sellado de lucernarios",
                importes={"MEDIO": Decimal("38000"), "LARGO": Decimal("38000")},
            )
        ],
    )
    fila = POR_CODIGO["HC.H02"].primera
    hoja = load_workbook(BytesIO(datos))["CapEx"]
    assert hoja[f"J{fila}"].value is None  # corto plazo vacío
    assert hoja[f"K{fila}"].value == 38000
    assert hoja[f"L{fila}"].value == 38000


def test_la_cabecera_del_encargo_llega_a_la_hoja_de_datos(encargo: Encargo) -> None:
    libro = load_workbook(BytesIO(generar(encargo, [_actuacion()])))
    activo = libro.worksheets[2]
    assert activo["C5"].value == "Encargo de prueba"
    assert activo["C8"].value == 1998
    assert activo["C16"].value == "INDUSTRIAL"


def test_una_categoria_que_la_plantilla_no_tiene_revienta(encargo: Encargo) -> None:
    """Se prefiere reventar a colocarla en el bloque de al lado: una actuación
    en la categoría equivocada suma mal y no se nota."""
    with pytest.raises(CeldaInexistente):
        generar(encargo, [_actuacion("HC.H99")])


# ─────────────────────────────────────────────────────────────────────────────
#  Cabida
# ─────────────────────────────────────────────────────────────────────────────


def test_la_plantilla_admite_diez_por_categoria() -> None:
    assert all(b.cabida == 10 for b in GEOMETRIA)


def test_once_actuaciones_no_caben_y_se_dice_cual(encargo: Encargo) -> None:
    """`[LIM]` Nunca se descarta una actuación en silencio: una que desaparece
    de la hoja que se manda al cliente es el fallo que nadie ve hasta que
    alguien suma a mano."""
    once = [_actuacion("HC.H08", descripcion=f"Actuación {i}") for i in range(11)]
    sobra = comprobar_cabida(once)
    assert [(d.categoria, d.hay, d.caben, d.sobran) for d in sobra] == [("HC.H08", 11, 10, 1)]
    with pytest.raises(NoCabe, match="HC.H08: 11 de 10"):
        generar(encargo, once)


def test_diez_caben_justas(encargo: Encargo) -> None:
    diez = [_actuacion("HC.H08", descripcion=f"Actuación {i}") for i in range(10)]
    assert comprobar_cabida(diez) == []
    hoja = load_workbook(BytesIO(generar(encargo, diez)))["CapEx"]
    assert hoja[f"G{POR_CODIGO['HC.H08'].ultima}"].value == "Actuación 9"


# ─────────────────────────────────────────────────────────────────────────────
#  Idioma
# ─────────────────────────────────────────────────────────────────────────────


def test_el_idioma_elige_la_plantilla_y_sus_etiquetas(encargo: Encargo) -> None:
    """La plantilla es la fuente de verdad de sus etiquetas: escribir «Cuartos
    Técnicos» donde el desplegable inglés espera «Technical Rooms» no da error,
    da una hoja que se abre bien y con los gráficos vacíos."""
    filas = {}
    for idioma in ("es", "en"):
        v = leer(idioma)
        datos = generar(
            encargo,
            [
                Actuacion(
                    categoria="HC.H09",
                    objeto=v.objeto("HC.H09.02"),
                    zona=v.zona("CUARTOS_TECNICOS", "INDUSTRIAL"),
                    descripcion="Cuadro general sin diferencial",
                    riesgo=v.riesgo("04"),
                    concepto=v.concepto("SEGURIDAD"),
                    recuperable=v.recuperable("SI"),
                    importes={"CORTO": Decimal("18400")},
                )
            ],
            idioma=idioma,
        )
        fila = POR_CODIGO["HC.H09"].primera
        hoja = load_workbook(BytesIO(datos))["CapEx"]
        filas[idioma] = tuple(hoja[f"{c}{fila}"].value for c in ("E", "F", "H", "P", "Q"))

    assert filas["es"] == ("CGBT", "Cuartos Técnicos", "04 Extremo", "e. Seguridad", "SI")
    assert filas["en"] == (
        "Main LV Switchboard",
        "Technical Rooms",
        "04 Extreme",
        "e. Safety",
        "YES",
    )


def test_un_idioma_sin_plantilla_se_rechaza(encargo: Encargo) -> None:
    with pytest.raises(ValueError, match="idioma no soportado"):
        generar(encargo, [_actuacion()], idioma="fr")


# ─────────────────────────────────────────────────────────────────────────────
#  Que la plantilla inglesa siga en inglés
# ─────────────────────────────────────────────────────────────────────────────
#
# La inglesa se hizo copiando la española y quedaron cinco textos sin traducir,
# repartidos por once celdas. `tools/traducir_plantilla_capex.py` los corrigió.
# Esta prueba impide que vuelvan a colarse si alguien actualiza la plantilla.


def test_la_plantilla_inglesa_no_tiene_texto_en_espanol() -> None:
    import re
    import zipfile

    with zipfile.ZipFile(PLANTILLAS / FICHERO["en"]) as z:
        cadenas = z.read("xl/sharedStrings.xml").decode("utf-8")
    visibles = re.findall(r"<t[^>]*>([^<]{3,})</t>", cadenas)
    # Palabras que no existen en inglés y sí en el original español.
    delatoras = re.compile(
        r"\b(procede|necesario|eliminar|adaptar|fuera|imprevistos|otras|otros|"
        r"licencias|obras|tasas|construcción|estimado)\b",
        re.IGNORECASE,
    )
    assert [t for t in visibles if delatoras.search(t)] == []


def test_la_categoria_h15_coincide_con_su_bloque_en_las_dos_plantillas() -> None:
    """El desplegable de categorías ofrecía «H15.Otros» mientras la cabecera
    del bloque decía «H15.Others»: el valor no estaba en su propia lista."""
    for idioma in IDIOMAS:
        libro = load_workbook(PLANTILLAS / FICHERO[idioma], keep_vba=True)
        categoria = libro.worksheets[0]["C18"].value
        cabecera = libro["CapEx"][f"A{POR_CODIGO['HC.H15'].subtotal}"].value
        if isinstance(cabecera, str) and not cabecera.startswith("="):
            assert categoria == cabecera, f"{idioma}: {categoria!r} != {cabecera!r}"
        libro.close()


# ─────────────────────────────────────────────────────────────────────────────
#  El puente entre el catálogo y la plantilla
# ─────────────────────────────────────────────────────────────────────────────


def test_todo_el_catalogo_tiene_etiqueta_en_las_dos_plantillas() -> None:
    """Si el catálogo crece y la plantilla no, la exportación escribiría un
    valor fuera de la lista del desplegable: la hoja se abriría bien y las
    tablas dinámicas dejarían esa fila fuera, sin decir nada."""
    from tdd.exports.vocabulario_capex import _casillas_de_objeto

    codigos = set(_casillas_de_objeto())
    assert codigos, "no se ha construido el puente catálogo → plantilla"
    for idioma in IDIOMAS:
        faltan = [c for c in codigos if leer(idioma).objetos.get(c) is None]
        assert faltan == [], f"{idioma}: sin etiqueta para {faltan}"


def test_el_general_de_medioambiental_no_se_empareja_por_posicion() -> None:
    """El caso que rompió el emparejamiento posicional: el catálogo conserva
    `General` el primero para no renumerar lo ya codificado, y la plantilla lo
    pone el último de su lista. Por índice, `MA.General.01` habría salido
    escrito como «Situación legal»."""
    assert leer("es").objeto("MA.General.01") == "General"
    assert leer("en").objeto("MA.General.01") == "General"
    assert leer("es").objeto("MA.General.02") == "Situación legal"
    assert leer("en").objeto("MA.General.02") == "Legal status"
