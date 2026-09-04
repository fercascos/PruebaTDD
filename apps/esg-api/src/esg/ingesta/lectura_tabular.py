"""Leer la tabla que venga: CSV o XLSX.

Este módulo solo saca **cabeceras y filas de texto**. No sabe qué es un
consumo. Esa separación es la que permite probar el diablo de los ficheros
reales —el punto y coma, el BOM de Excel, la coma decimal, la fecha
dd/mm/aaaa— sin montar nada más.
"""

from __future__ import annotations

import csv
import hashlib
import io
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any


class FicheroIlegible(Exception):
    """Ni CSV ni XLSX, o roto. Se dice cuál de las dos cosas."""


@dataclass(frozen=True, slots=True)
class Tabla:
    cabeceras: list[str]
    #: Cada fila ya emparejada con su número REAL en el fichero (la 1 es la de
    #: cabeceras), porque «error en la fila 143» tiene que poder buscarse en el
    #: Excel del cliente sin restar mentalmente.
    filas: list[tuple[int, dict[str, Any]]]
    hoja: str | None = None


def sha256(contenido: bytes) -> str:
    return hashlib.sha256(contenido).hexdigest()


def _decodificar(contenido: bytes) -> str:
    """UTF-8 y, si no, la codificación que suelta Excel en español.

    No se adivina con heurísticas: se prueban dos y se falla con un mensaje
    claro. Un fichero mal decodificado no da error, da acentos rotos en los
    nombres de los activos, y eso se descubre tres semanas después.
    """
    for codec in ("utf-8-sig", "cp1252"):
        try:
            return contenido.decode(codec)
        except UnicodeDecodeError:
            continue
    raise FicheroIlegible("El fichero no está en UTF-8 ni en Windows-1252")


def leer_csv(contenido: bytes) -> Tabla:
    texto = _decodificar(contenido)
    muestra = texto[:4096]
    try:
        # `;` primero en la lista de candidatos: es lo que produce Excel con
        # configuración regional española, que es de donde vienen estos
        # ficheros. Sin acotar los candidatos, el Sniffer elige a veces el
        # espacio y la tabla sale de una sola columna.
        dialecto: Any = csv.Sniffer().sniff(muestra, delimiters=";,\t|")
    except csv.Error:
        dialecto = csv.excel
        dialecto.delimiter = ";" if muestra.count(";") > muestra.count(",") else ","
    lector = csv.DictReader(io.StringIO(texto), dialect=dialecto)
    cabeceras = [c.strip() for c in (lector.fieldnames or [])]
    if not cabeceras or cabeceras == [""]:
        raise FicheroIlegible("El fichero no tiene fila de cabeceras")
    filas: list[tuple[int, dict[str, Any]]] = []
    for numero, fila in enumerate(lector, start=2):
        limpia = {(k or "").strip(): v for k, v in fila.items()}
        if all(_vacio(v) for v in limpia.values()):
            continue  # fila en blanco: no es un error, es un Excel
        filas.append((numero, limpia))
    return Tabla(cabeceras=cabeceras, filas=filas)


def leer_xlsx(contenido: bytes, *, hoja: str | None = None) -> Tabla:
    from openpyxl import load_workbook

    try:
        libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception as exc:
        raise FicheroIlegible("El fichero no es un XLSX legible") from exc
    try:
        pagina = libro[hoja] if hoja else libro[libro.sheetnames[0]]
    except KeyError as exc:
        raise FicheroIlegible(
            f"El libro no tiene una hoja llamada «{hoja}». Tiene: " + ", ".join(libro.sheetnames)
        ) from exc

    iterador = pagina.iter_rows(values_only=True)
    try:
        cabecera_bruta = next(iterador)
    except StopIteration as exc:
        raise FicheroIlegible("La hoja está vacía") from exc
    cabeceras = [str(c).strip() if c is not None else "" for c in cabecera_bruta]
    filas: list[tuple[int, dict[str, Any]]] = []
    for numero, valores in enumerate(iterador, start=2):
        if all(_vacio(v) for v in valores):
            continue
        filas.append((numero, dict(zip(cabeceras, valores, strict=False))))
    libro.close()
    return Tabla(cabeceras=cabeceras, filas=filas, hoja=pagina.title)


def leer(contenido: bytes, *, nombre: str, hoja: str | None = None) -> Tabla:
    """Elige lector por el nombre del fichero, y comprueba la firma.

    `[REQ]` Un XLSX renombrado a `.csv` —pasa más de lo que parece— se lee como
    texto binario y produce cien incidencias sin sentido en vez de una frase
    útil. Los dos primeros bytes de un XLSX son los de un ZIP.
    """
    minusculas = nombre.lower()
    parece_zip = contenido[:2] == b"PK"
    if minusculas.endswith((".xlsx", ".xlsm")) or parece_zip:
        if not parece_zip:
            raise FicheroIlegible("El fichero tiene extensión de Excel pero no es un XLSX")
        return leer_xlsx(contenido, hoja=hoja)
    if minusculas.endswith((".csv", ".txt", ".tsv")):
        return leer_csv(contenido)
    raise FicheroIlegible(f"No se sabe leer «{nombre}»: se admiten CSV y XLSX")


def _vacio(v: object) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def texto_de(valor: object) -> str:
    """El valor de una celda como texto, sin sorpresas de openpyxl.

    Excel devuelve `datetime` en las fechas y `float` en los números; pasar
    `str()` a un `datetime` da «2025-03-01 00:00:00», que luego no casa con
    ningún formato de fecha. Cada tipo se convierte a lo que se espera de él.
    """
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()
